from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font

# Function to format the transaction details
def format_transaction_details(transaction):
    customer = transaction['customer']
    authorization = transaction['authorization']

    formatted_details = f"""
    Transaction Statement:
    -----------------------
    Transaction ID: {transaction['id']}
    Status: {transaction['status'].capitalize()}
    Amount: {transaction['currency']} {transaction['amount'] / 100:.2f}
    Message: {transaction['message']}
    Date Created: {transaction['created_at']}
    Payment Channel: {transaction['channel'].capitalize()}
    Bank: {authorization['bank']}
    Authorization Code: {authorization['authorization_code']}
    Requested Amount: {transaction['currency']} {transaction['requested_amount'] / 100:.2f}
    """

    return formatted_details




def format_transactions_to_excel(transactions, file_name='transactions_details.xlsx'):
    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Transaction Details'

    # Define the header row
    headers = [
        'Transaction ID', 'Status', 'Amount', 'Currency', 'Message',
        'Date Created', 'Payment Channel', 'Bank', 'Authorization Code', 'Requested Amount'
    ]
    
    # Write the header row with formatting
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)
        sheet.cell(row=1, column=col_num).font = Font(bold=True)
        sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

    # Write the transaction details into rows
    for row_num, transaction in enumerate(transactions, start=2):
        authorization = transaction.get('authorization', {})
        created_at = transaction.get('created_at', 'N/A')
        try:
            created_at = datetime.strptime(created_at, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%d %b %Y %H:%M:%S')
        except ValueError:
            pass
        
        details = [
            transaction.get('id', 'N/A'),
            transaction.get('status', 'unknown').capitalize(),
            transaction.get('amount', 0) / 100,
            transaction.get('currency', 'N/A'),
            transaction.get('message', 'No message'),
            created_at,
            transaction.get('channel', 'unknown').capitalize(),
            authorization.get('bank', 'N/A'),
            authorization.get('authorization_code', 'N/A'),
            transaction.get('requested_amount', 0) / 100
        ]
        
        for col_num, value in enumerate(details, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)
            sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left')

    # Adjust column widths
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save the Excel file
    workbook.save(file_name)

    print(f"Transaction details saved to {file_name}")

    return file_name
